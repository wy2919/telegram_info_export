import re
import socks
import traceback
import pandas as pd
from openpyxl import load_workbook
from telethon import errors
from telethon.sync import TelegramClient
from telethon.tl.types import User
from openpyxl.worksheet.table import Table, TableStyleInfo


def get_telegram_info(api_name, api_id, api_hash, phone_number, proxy_url):
    try:
        if proxy_url:
            print(f"当前代理为: {proxy_url}")
            proxy = get_proxy(proxy_url)
        else:
            print("当前没有设置代理")
            proxy = None

        print("正在创建 Telegram 客户端")
        client = TelegramClient(
            api_name,
            int(api_id),
            api_hash,
            proxy=proxy,
            connection_retries=5,
        )
        print("正在连接到 Telegram 服务器")
        client.connect()
        print("连接成功")

        if not client.is_user_authorized():
            print("当前账号需要验证码")
            client.send_code_request(phone_number)

            verification_code = input("请输入验证码: ")

            try:
                client.sign_in(phone_number, verification_code)
            except errors.SessionPasswordNeededError:
                print("当前账号需要二步验证")

                password = input("请输入二步验证密码: ")

                client.sign_in(password=password)

            print("登陆成功")

        print("正在获取对话信息")

        groups_and_channels = []
        bots = []
        contact_person = []

        for dialog in client.iter_dialogs():
            if dialog.is_group:
                groups_and_channels.append((dialog.entity, "群组", dialog.name))
            elif dialog.is_channel:
                groups_and_channels.append((dialog.entity, "频道", dialog.name))
            elif dialog.entity.bot:
                bots.append(
                    (dialog.entity.username, dialog.entity.first_name, dialog.entity.id)
                )
            elif isinstance(dialog.entity, User) and not dialog.entity.bot:
                username = dialog.entity.username
                first_name = dialog.entity.first_name or ""
                last_name = dialog.entity.last_name or ""
                full_name = f"{first_name} {last_name}".strip()
                if username:
                    contact_person.append((username, full_name, dialog.entity.id))

        def get_display_width(s):
            if not isinstance(s, str):
                s = str(s)
            zh_pattern = re.compile("[\u4e00-\u9fa5]")
            width = 0
            for ch in s:
                if zh_pattern.match(ch):
                    width += 2
                else:
                    width += 1
            return width

        group_channel_with_links = []
        group_channel_without_links = []

        for entity, chat_type, chat_title in groups_and_channels:
            username = getattr(entity, "username", None)
            if username:
                invite_link = f"https://t.me/{username}"
                group_channel_with_links.append(
                    {"类型": chat_type, "名称": chat_title, "邀请链接": invite_link}
                )
            else:
                group_channel_without_links.append(
                    {"类型": chat_type, "名称": chat_title, "邀请链接": ""}
                )

        bot_data = []
        for username, first_name, user_id in bots:
            bot_data.append({"用户名": f"@{username}", "昵称": first_name})

        contact_person_data = []
        for username, first_name, user_id in contact_person:
            contact_person_data.append({"用户名": f"@{username}", "昵称": first_name})

        # 创建DataFrame
        df_groups_channels = pd.DataFrame(group_channel_with_links)
        df_groups_channels_no_invite = pd.DataFrame(group_channel_without_links)
        df_bots = pd.DataFrame(bot_data)
        df_contact_person = pd.DataFrame(contact_person_data)

        file_path = api_id+"_telegram_info.xlsx"
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_groups_channels.to_excel(writer, sheet_name="群组和频道", index=False)
            df_groups_channels_no_invite.to_excel(
                writer, sheet_name="群组和频道(无邀请)", index=False
            )
            df_bots.to_excel(writer, sheet_name="机器人", index=False)
            df_contact_person.to_excel(writer, sheet_name="联系人", index=False)

        print("正在导出")

        # 加载Excel文件并添加筛选功能
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 创建表格
            tab = Table(displayName=f"Table_{sheet_name}", ref=ws.dimensions)
            # 添加样式
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        length = get_display_width(cell.value)
                        if length > max_length:
                            max_length = length
                    except (
                            AttributeError,
                            TypeError,
                            ValueError,
                            UnicodeEncodeError,
                            UnicodeDecodeError,
                    ):
                        pass
                adjusted_width = max_length + 4
                ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)

        print(f"导出成功, 请查看{file_path}文件")

    except Exception as e:
        print(f"发生错误: {str(e)}")
        print(f"错误详情: {traceback.format_exc()}")
    finally:
        if "client" in locals() and client.is_connected():
            client.disconnect()


def get_proxy(proxy_url):
    re_pattern = r"^(?P<protocol>http|socks5)://(?:(?P<user>[^:]+):(?P<password>[^@]+)@)?(?P<host>[^:]+):(?P<port>\d+)$"
    match = re.match(re_pattern, proxy_url)
    if not match:
        return None

    proxy_protocol = match.group("protocol")
    proxy_host = match.group("host")
    proxy_port = int(match.group("port"))
    proxy_user = match.group("user")
    proxy_pass = match.group("password")

    proxy_type = socks.SOCKS5 if proxy_protocol == "socks5" else socks.HTTP
    return (
        proxy_type,
        proxy_host,
        proxy_port,
        bool(proxy_user),
        proxy_user,
        proxy_pass,
    )


if __name__ == "__main__":
    api_id = "15xxxxx"
    api_hash = "543xxxxxxxxxxxxxxxxxxx"
    phone_number = "+130xxxxx"
    proxy_url = "socks5://seeve:20080"

    get_telegram_info("anon", api_id, api_hash,phone_number,proxy_url)
