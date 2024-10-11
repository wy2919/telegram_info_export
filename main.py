import re
import os
import socks
import configparser
import traceback
import pandas as pd
import tkinter as tk
from ttkthemes import ThemedTk
from openpyxl import load_workbook
from telethon.sync import TelegramClient
from telethon.tl.types import User
from tkinter import ttk, simpledialog
from openpyxl.worksheet.table import Table, TableStyleInfo


def get_telegram_info(self, api_name, api_id, api_hash, phone_number, proxy_url):
    try:
        if proxy_url:
            self.log(f"当前代理为: {proxy_url}")
            proxy = get_proxy(proxy_url)
        else:
            self.log("当前没有设置代理")
            proxy = None

        self.log("正在创建 Telegram 客户端")
        client = TelegramClient(
            api_name,
            int(api_id),
            api_hash,
            proxy=proxy,
            connection_retries=5,
        )
        self.log("正在连接到 Telegram 服务器")
        client.connect()
        self.log("连接成功")

        if not client.is_user_authorized():
            self.log("当前账号需要验证")
            client.send_code_request(phone_number)
            verification_code = simpledialog.askstring(
                title="验证码", prompt="请输入验证码:"
            )
            client.sign_in(phone_number, verification_code)
            self.log("登陆成功")

        self.log("正在获取对话信息")

        groups_and_channels = []
        bots = []
        contact_person = []

        for dialog in client.iter_dialogs():
            if dialog.is_group:
                groups_and_channels.append((dialog.entity, "群组", dialog.name))
            elif dialog.is_channel:
                groups_and_channels.append((dialog.entity, "频道", dialog.name))
            elif dialog.entity.bot:
                bots.append(
                    (dialog.entity.username, dialog.entity.first_name, dialog.entity.id)
                )
            elif isinstance(dialog.entity, User) and not dialog.entity.bot:
                username = dialog.entity.username
                first_name = dialog.entity.first_name or ""
                last_name = dialog.entity.last_name or ""
                full_name = f"{first_name} {last_name}".strip()
                if username:
                    contact_person.append((username, full_name, dialog.entity.id))

        def get_display_width(s):
            if not isinstance(s, str):
                s = str(s)
            zh_pattern = re.compile("[\u4e00-\u9fa5]")
            width = 0
            for ch in s:
                if zh_pattern.match(ch):
                    width += 2
                else:
                    width += 1
            return width

        group_channel_with_links = []
        group_channel_without_links = []

        for entity, chat_type, chat_title in groups_and_channels:
            username = getattr(entity, "username", None)
            if username:
                invite_link = f"https://t.me/{username}"
                group_channel_with_links.append(
                    {"类型": chat_type, "名称": chat_title, "邀请链接": invite_link}
                )
            else:
                group_channel_without_links.append(
                    {"类型": chat_type, "名称": chat_title, "邀请链接": ""}
                )

        bot_data = []
        for username, first_name, user_id in bots:
            bot_data.append({"用户名": f"@{username}", "昵称": first_name})

        contact_person_data = []
        for username, first_name, user_id in contact_person:
            contact_person_data.append({"用户名": f"@{username}", "昵称": first_name})

        # 创建DataFrame
        df_groups_channels = pd.DataFrame(group_channel_with_links)
        df_groups_channels_no_invite = pd.DataFrame(group_channel_without_links)
        df_bots = pd.DataFrame(bot_data)
        df_contact_person = pd.DataFrame(contact_person_data)

        file_path = "telegram_info.xlsx"
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_groups_channels.to_excel(writer, sheet_name="群组和频道", index=False)
            df_groups_channels_no_invite.to_excel(
                writer, sheet_name="群组和频道(无邀请)", index=False
            )
            df_bots.to_excel(writer, sheet_name="机器人", index=False)
            df_contact_person.to_excel(writer, sheet_name="联系人", index=False)

        self.log("正在导出")

        # 加载Excel文件并添加筛选功能
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 创建表格
            tab = Table(displayName=f"Table_{sheet_name}", ref=ws.dimensions)
            # 添加样式
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        length = get_display_width(cell.value)
                        if length > max_length:
                            max_length = length
                    except (
                        AttributeError,
                        TypeError,
                        ValueError,
                        UnicodeEncodeError,
                        UnicodeDecodeError,
                    ):
                        pass
                adjusted_width = max_length + 4
                ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)

        self.log("导出成功, 请查看telegram_info.xlsx文件")

    except Exception as e:
        self.log(f"发生错误: {str(e)}")
        self.log(f"错误详情: {traceback.format_exc()}")
    finally:
        if "client" in locals() and client.is_connected():
            client.disconnect()


def get_proxy(proxy_url):
    re_pattern = r"^(?P<protocol>http|socks5)://(?:(?P<user>[^:]+):(?P<password>[^@]+)@)?(?P<host>[^:]+):(?P<port>\d+)$"
    match = re.match(re_pattern, proxy_url)
    if not match:
        return None

    proxy_protocol = match.group("protocol")
    proxy_host = match.group("host")
    proxy_port = int(match.group("port"))
    proxy_user = match.group("user")
    proxy_pass = match.group("password")

    proxy_type = socks.SOCKS5 if proxy_protocol == "socks5" else socks.HTTP
    return (
        proxy_type,
        proxy_host,
        proxy_port,
        bool(proxy_user),
        proxy_user,
        proxy_pass,
    )


class Main:
    def __init__(self):
        # 主界面
        self.root = ThemedTk(theme="yaru")
        self.root.title("导出Telegram信息")

        ttk.Label(self.root, text="api_id:").grid(row=0, column=0, padx=10, pady=5)
        ttk.Label(self.root, text="api_hash:").grid(row=1, column=0, padx=10, pady=5)
        ttk.Label(self.root, text="phone_number:").grid(
            row=2, column=0, padx=10, pady=5
        )
        ttk.Label(self.root, text="proxy_url:").grid(row=3, column=0, padx=10, pady=5)

        # 创建输入框
        self.api_id = ttk.Entry(self.root, width=36)
        self.api_hash = ttk.Entry(self.root, width=36)
        self.phone_number = ttk.Entry(self.root, width=36)
        self.proxy_url = ttk.Entry(self.root, width=36)

        self.api_id.grid(row=0, column=1, padx=10, pady=5)
        self.api_hash.grid(row=1, column=1, padx=10, pady=5)
        self.phone_number.grid(row=2, column=1, padx=10, pady=5)
        self.proxy_url.grid(row=3, column=1, padx=10, pady=5)

        button_frame = ttk.Frame(self.root)
        button_frame.grid(row=4, column=0, columnspan=2, pady=10)

        ttk.Button(
            button_frame,
            text="开始",
            command=lambda: get_telegram_info(
                self,
                "anon",
                self.api_id.get(),
                self.api_hash.get(),
                self.phone_number.get(),
                self.proxy_url.get(),
            ),
        ).pack(side="left", padx=10)

        self.log_text = tk.Text(self.root, height=10, width=50)
        self.log_text.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

        self.load_config()
        self.log("点击开始进行导出Telegram信息，请耐心等待")

        self.root.mainloop()

    def load_config(self):
        config = configparser.ConfigParser()
        if not os.path.isfile("config.ini"):
            self.log("配置文件不存在，请手动填写。")
        else:
            try:
                config.read("config.ini")
                missing_values = []

                for key in ["api_id", "api_hash", "phone_number", "proxy_url"]:
                    value = config["DEFAULT"].get(key, "").strip()
                    if not value:
                        missing_values.append(key)
                    else:
                        getattr(self, key).insert(0, value)

                if missing_values:
                    self.log(f"配置文件中{', '.join(missing_values)}为空，请手动补全")
            except KeyError:
                self.log("配置文件中缺少环境变量，请手动补全")

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)


if __name__ == "__main__":
    Main()
